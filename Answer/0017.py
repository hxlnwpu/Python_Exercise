#coding='utf-8'

from openpyxl import load_workbook
import xml.etree.ElementTree as ET
from xml.dom import minidom

wb = load_workbook(r"C:\Workspace\Python\Python_Exercise\Data\student.xlsx")
path=r"C:\Workspace\Python\Python_Exercise\Data\student.xml"
ws=wb['Sheet1']
data=ws['A1':'E3']
L=[]
i=0
L.append("<!--"+'\n'+"学生信息表"+'\n'+"id"+": [名字, 数学, 语文, 英文] -->"+'\n')
L.append('{'+'\n')
for line in data:
    if(i<2):
        line='"'+str(line[0].value)+'"'+": ["+'"'+str(line[1].value)+'"'+','+str(line[2].value)+','+str(line[3].value)+','+str(line[4].value)+']'+','+'\n'
        L.append(line)
        i=i+1
    else:
        line='"'+str(line[0].value)+'"'+": ["+'"'+str(line[1].value)+'"'+','+str(line[2].value)+','+str(line[3].value)+','+str(line[4].value)+']'+'\n'
        L.append(line)
L.append('}'+'\n')
root=ET.Element('root')
students=ET.SubElement(root,"students")
studentText=L[0]+L[1]+L[2]+L[3]+L[4]+L[5]
students.text=studentText
tree=ET.ElementTree(root)
tree.write('../Data/students.xml',encoding='utf-8')
